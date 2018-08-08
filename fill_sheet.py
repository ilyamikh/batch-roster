from openpyxl import load_workbook
from openpyxl import styles
import parse_roster
import os
import time
import datetime
from math import ceil


def process_list(file='raw.xml'):
    """
    Takes the path of the raw XML file and process the roster to generate classroom lists
    :param file: path to raw XML data
    :return: none
    """

    print("Starting process...")
    rles = parse_roster.get_roster(file)

    output_choice = None
    while output_choice != 'q':
        output_choice = input("Enter 'm' for Meal Count or 'r' for Roster. 'q' to quit.")

        if output_choice == 'r':
            user_month = input("Month/Year? (Leave blank to use current date)")
            if user_month:
                month = user_month
            else:
                month = datetime.datetime.now().strftime('%B, %Y')
            filepath = "Roster_" + get_timestamps()
            if not os.path.exists(filepath):
                os.makedirs(filepath)
            create_monthly_rosters(rles, filepath, month)

        elif output_choice == 'm':
            user_date = input("Week of MM-DD-YYYY? (Leave blank to use current date)")
            if user_date:
                date = user_date
            else:
                date = get_current_date()
            filepath = "Meal_Count_" + get_timestamps()
            if not os.path.exists(filepath):
                os.makedirs(filepath)
            create_meal_rosters(rles, filepath, date)


def create_meal_rosters(rles, filepath, month):
    for group in rles:
        classroom = rles[group]  # they're sorted in make_meal_sheet, is that inconsistent?
        children = len(classroom)
        print(group, children, "children.")
        make_meal_sheet(group, filepath, classroom, month)

    print("Processed", len(rles), "groups.")


def create_monthly_rosters(rles, filepath, month):
    for group in rles:  # I'll try and take care of the logic to split the sheets here
        classroom = sorted(rles[group])  # sort by name takes place here
        children = len(classroom)
        print(group, children, "children.")
        if children < 14:
            make_roster_sheet(group, filepath, classroom, month)  # group is the key or name of the class
        else:
            for i in range(0, ceil(children / 13)):
                group_name = group + "_" + str(i+1)

                start = i*13
                stop = i*13 + 13
                if stop > children:
                    stop = children  # what't the one line way to do this?

                make_roster_sheet(group_name, filepath, classroom[start:stop], month)
    print("Processed", len(rles), "groups.")


def get_timestamps():
    """Returns simply formatted timestamp string"""
    return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d-%H-%M-%S')


def get_current_date():
    return datetime.datetime.fromtimestamp(time.time()).strftime('%m-%d-%Y')


def get_date_obj(date_string):
    """Takes a string in the format mm-dd-yyyy and returns a datetime object."""
    return datetime.datetime.strptime(date_string, '%m-%d-%Y')


def format_date(date_obj):
    """Takes a datetime object and returns a mm-dd-yyyy format string"""
    return datetime.datetime.strftime(date_obj, '%m-%d-%Y')


def get_month(date_obj):
    return datetime.datetime.strftime(date_obj, '%B')


def get_year(date_obj):
    return datetime.datetime.strftime(date_obj, '%Y')


def make_meal_sheet(group, filepath, classroom, start_date):
    """Processes each group, creating and saving the meal count sheet. Maximum length is 33 children."""
    print("Filling", group + ', ', len(classroom), "children...")

    book = load_workbook("meal_count_template.xlsx")
    name = filepath + "/" + group.replace('/', '') + ".xlsx"
    # date variables go here
    sheet = book.active

    sheet['E7'] = format_date(get_date_obj(start_date))
    sheet['I7'] = format_date(get_date_obj(start_date) + datetime.timedelta(days=1))
    sheet['M7'] = format_date(get_date_obj(start_date) + datetime.timedelta(days=2))
    sheet['Q7'] = format_date(get_date_obj(start_date) + datetime.timedelta(days=3))
    sheet['U7'] = format_date(get_date_obj(start_date) + datetime.timedelta(days=4))

    sheet['B50'] = "Categories current as of " + get_timestamps()

    sheet['B6'] = "GROUP: " + group
    sheet['B7'] = "MONTH: " + get_month(get_date_obj(start_date))
    sheet['B8'] = "YEAR: " + get_year(get_date_obj(start_date))

    current_cell = 10  # see template
    current_cat = 1  # to add a blank line
    cat_count = dict()

    classroom.sort(key=lambda x: x[1])
    for child in classroom:
        cat_count[child[1]] = cat_count.get(child[1], 0) + 1  # simple tally
        if current_cat != child[1]:
            set_border(sheet, current_cell)
            current_cell += 1  # skip a line to separate the categories visually
            current_cat = child[1]
        cellname = 'B' + str(current_cell)
        sheet[cellname] = child[0]  # name
        cellname = 'C' + str(current_cell)
        sheet[cellname] = child[1]  # category
        current_cell += 1

    if 1 in cat_count:
        sheet['B47'] = "Category 1: " + str(cat_count[1])
    else:
        sheet['B47'] = "Category 1: 0"

    if 2 in cat_count:
        sheet['B48'] = "Category 2: " + str(cat_count[2])
    else:
        sheet['B48'] = "Category 2: 0"

    if 3 in cat_count:
        sheet['B49'] = "Category 3: " + str(cat_count[3])
    else:
        sheet['B49'] = "Category 3: 0"

    book.save(name)
    book.close()
    print("Done")


def set_border(sheet, row):
    thick_bottom = styles.Border(bottom=styles.Side(style='thick'))
    col = 'B'
    while col != 'X':
        cell = col + str(row)
        sheet[cell].border = thick_bottom
        col = chr(ord(col) + 1)  # the alphabet is indeed in order


def make_roster_sheet(group, filepath, classroom, in_month):
    """
    Processes each group, creating and saving the roster
    :param group: name of the group
    :param filepath: path to save the output
    :param classroom: list of children
    :return: none
    """
    print("Filling", group + ', ', len(classroom), "children...")

    book = load_workbook("roster_template.xlsx")
    name = filepath + "/" + group.replace('/', '') + ".xlsx"
    month = in_month
    sheet = book.active

    sheet['B1'] = group
    sheet['Z1'] = month

    current_cell = 5  # see template
    for child in classroom:
        cellname = 'A' + str(current_cell)
        sheet[cellname] = child[0]
        current_cell += 2

    book.save(name)
    book.close()
    print("Done")

